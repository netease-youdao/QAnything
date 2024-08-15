from qanything_kernel.utils.general_utils import get_time, get_table_infos, num_tokens_embed, get_all_subpages, \
    html_to_markdown
from typing import List, Optional
from qanything_kernel.configs.model_config import UPLOAD_ROOT_PATH, LOCAL_OCR_SERVICE_URL, IMAGES_ROOT_PATH, \
    DEFAULT_CHILD_CHUNK_SIZE, LOCAL_PDF_PARSER_SERVICE_URL
from langchain.docstore.document import Document
from qanything_kernel.utils.loader.my_recursive_url_loader import MyRecursiveUrlLoader
from qanything_kernel.utils.custom_log import insert_logger
from langchain_community.document_loaders import UnstructuredFileLoader, TextLoader
from langchain_community.document_loaders import UnstructuredWordDocumentLoader
from langchain_community.document_loaders import UnstructuredEmailLoader
from langchain_community.document_loaders import UnstructuredPowerPointLoader
from qanything_kernel.utils.loader import UnstructuredPaddlePDFLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter
from qanything_kernel.utils.loader.csv_loader import CSVLoader
from qanything_kernel.utils.loader.markdown_parser import convert_markdown_to_langchaindoc
import asyncio
import aiohttp
import docx2txt
import base64
import pandas as pd
import os
import json
import requests
import threading
import re
import newspaper
import uuid
import traceback
import openpyxl
import shutil


def get_ocr_result_sync(image_data):
    try:
        response = requests.post(f"http://{LOCAL_OCR_SERVICE_URL}/ocr", data=image_data)
        response.raise_for_status()  # 如果请求返回了错误状态码，将会抛出异常
        ocr_res = response.text
        ocr_res = json.loads(ocr_res)
        return ocr_res['result']
    except Exception as e:
        insert_logger.warning(f"ocr error: {traceback.format_exc()}")
        return None


class LocalFileForInsert:
    def __init__(self, user_id, kb_id, file_id, file_location, file_name, file_url, chunk_size, mysql_client):
        self.chunk_size = chunk_size
        self.markdown_text_splitter = RecursiveCharacterTextSplitter(chunk_size=chunk_size, chunk_overlap=0,
                                                                     length_function=num_tokens_embed)
        self.user_id = user_id
        self.kb_id = kb_id
        self.file_id = file_id
        self.docs: List[Document] = []
        self.embs = []
        self.file_name = file_name
        self.file_location = file_location
        self.file_url = ""
        self.faq_dict = {}
        self.file_path = ""
        self.mysql_client = mysql_client
        if self.file_location == 'FAQ':
            faq_info = self.mysql_client.get_faq(self.file_id)
            user_id, kb_id, question, answer, nos_keys = faq_info
            self.faq_dict = {'question': question, 'answer': answer, 'nos_keys': nos_keys}
        elif self.file_location == 'URL':
            self.file_url = file_url
            upload_path = os.path.join(UPLOAD_ROOT_PATH, user_id)
            file_dir = os.path.join(upload_path, self.kb_id, self.file_id)
            os.makedirs(file_dir, exist_ok=True)
            self.file_path = os.path.join(file_dir, self.file_name)
        else:
            self.file_path = self.file_location
        self.event = threading.Event()

    @staticmethod
    @get_time
    def image_ocr_txt(filepath, dir_path="tmp_files"):
        full_dir_path = os.path.join(os.path.dirname(filepath), dir_path)
        if not os.path.exists(full_dir_path):
            os.makedirs(full_dir_path)
        filename = os.path.split(filepath)[-1]

        # 读取图片
        img_np = open(filepath, 'rb').read()

        img_data = {
            "img64": base64.b64encode(img_np).decode("utf-8"),
        }

        result = get_ocr_result_sync(img_data)

        ocr_result = [line for line in result if line]
        ocr_result = '\n'.join(ocr_result)

        insert_logger.info(f'ocr_res[:100]: {ocr_result[:100]}')

        # 写入结果到文本文件
        txt_file_path = os.path.join(full_dir_path, "%s.txt" % (filename))
        with open(txt_file_path, 'w', encoding='utf-8') as fout:
            fout.write(ocr_result)

        return txt_file_path

    def table_process(self, doc):
        table_infos = get_table_infos(doc.page_content)
        title_lst = doc.metadata['title_lst']
        new_docs = []
        if table_infos is not None:
            tmp_content = '\n'.join(title_lst) + '\n' + doc.page_content
            if num_tokens_embed(tmp_content) <= self.chunk_size:
                doc.page_content = tmp_content
                return [doc]
            head_line = table_infos['head_line']
            end_line = table_infos['end_line']
            table_head = table_infos['head']

            # 处理表格前的内容
            if head_line != 0:
                tmp_doc = Document(
                    page_content='\n'.join(title_lst) + '\n' + '\n'.join(table_infos['lines'][:head_line]),
                    metadata=doc.metadata)
                new_docs.append(tmp_doc)

            # 处理表格内容
            table_content = '\n'.join(title_lst) + '\n' + table_head
            for line in table_infos['lines'][head_line + 2:end_line + 1]:
                if num_tokens_embed(table_content + '\n' + line) > self.chunk_size:
                    # 如果添加新行会超出chunk_size，先保存当前内容
                    tmp_doc = Document(page_content=table_content, metadata=doc.metadata)
                    new_docs.append(tmp_doc)
                    # 重新开始一个新的chunk，包含标题和表头
                    table_content = '\n'.join(title_lst) + '\n' + table_head + '\n' + line
                else:
                    table_content += '\n' + line

            # 保存最后一个chunk
            if table_content != '\n'.join(title_lst) + '\n' + table_head:
                tmp_doc = Document(page_content=table_content, metadata=doc.metadata)
                new_docs.append(tmp_doc)

            # 处理表格后的内容
            if end_line != len(table_infos['lines']) - 1:
                tmp_doc = Document(
                    page_content='\n'.join(title_lst) + '\n' + '\n'.join(table_infos['lines'][end_line:]),
                    metadata=doc.metadata)
                new_docs.append(tmp_doc)

            insert_logger.info(f"TABLE SLICES: {new_docs[:2]}")
        else:
            return None
        return new_docs

    @staticmethod
    def get_page_id(doc, pre_page_id):
        # 查找 page_id 标志行
        lines = doc.page_content.split('\n')
        for line in lines:
            if re.match(r'^#+ 当前页数:\d+$', line):
                try:
                    page_id = int(line.split(':')[-1])
                    return page_id
                except ValueError:
                    continue
        return pre_page_id

    def markdown_process(self, docs: List[Document]):
        new_docs = []
        for doc in docs:
            title_lst = doc.metadata['title_lst']
            # 删除所有仅有多个#的title
            title_lst = [t for t in title_lst if t.replace('#', '') != '']
            has_table = doc.metadata['has_table']
            if has_table:
                table_doc_id = str(uuid.uuid4())
                self.mysql_client.add_document(table_doc_id, doc.to_json())
                doc.metadata['table_doc_id'] = table_doc_id
                table_docs = self.table_process(doc)
                if table_docs:
                    new_docs.extend(table_docs)
                    continue
            slices = self.markdown_text_splitter.split_documents([doc])
            # insert_logger.info(f"markdown_text_splitter: {len(slices)}")
            if len(slices) == 1:
                slices[0].page_content = '\n\n'.join(title_lst) + '\n\n' + slices[0].page_content
            else:
                for idx, slice in enumerate(slices):
                    slice.page_content = '\n\n'.join(
                        title_lst) + f'\n\n###### 第{idx + 1}段内容如下：\n' + slice.page_content
            new_docs.extend(slices)
        return new_docs

    @get_time
    async def url_to_documents(self, file_path, file_name, file_url, dir_path="tmp_files", max_retries=3):
        full_dir_path = os.path.join(os.path.dirname(file_path), dir_path)
        if not os.path.exists(full_dir_path):
            os.makedirs(full_dir_path)

        for attempt in range(max_retries):
            try:
                headers = {
                    "Accept": "application/json",
                    "X-Return-Format": "markdown",
                    "X-Timeout": "15",
                }
                async with aiohttp.ClientSession() as session:
                    async with session.get(f"https://r.jina.ai/{file_url}", headers=headers, timeout=30) as response:
                        jina_response = await response.json()
                        if jina_response['code'] == 200:
                            title = jina_response['data'].get('title', '')
                            markdown_str = jina_response['data'].get('content', '')
                            markdown_str = html_to_markdown(markdown_str)
                            md_file_path = os.path.join(full_dir_path, "%s.md" % (file_name))
                            with open(md_file_path, 'w', encoding='utf-8') as fout:
                                fout.write(markdown_str)
                            docs = convert_markdown_to_langchaindoc(md_file_path)
                            if title:
                                for doc in docs:
                                    doc.metadata['title'] = title
                            docs = self.markdown_process(docs)
                            return docs
                        else:
                            insert_logger.warning(f"jina get url warning: {file_url}, {jina_response}")
            except Exception as e:
                insert_logger.warning(f"jina get url error: {file_url}, {traceback.format_exc()}")

            if attempt < max_retries - 1:  # 如果不是最后一次尝试，等待30秒后重试
                await asyncio.sleep(30)

        return None

    @staticmethod
    def excel_to_markdown(file_path, markdown_path):
        basename = os.path.basename(file_path)
        markdown_file = os.path.join(markdown_path, basename.split('.')[0] + '.md')
        # 打开 Excel 文件
        workbook = openpyxl.load_workbook(file_path)

        with open(markdown_file, 'w', encoding='utf-8') as md_file:
            # 遍历所有的工作表
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # 添加 sheet 名称作为标题
                md_file.write(f"# {sheet_name}\n\n")

                for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                    # 将每行转换为 Markdown 表格行
                    markdown_row = '| ' + ' | '.join(str(cell) if cell is not None else '' for cell in row) + ' |'
                    md_file.write(markdown_row + '\n')

                    # 在第一行后添加分隔符
                    if row_index == 0:
                        separator = '|' + '|'.join(['---' for _ in row]) + '|'
                        md_file.write(separator + '\n')

                # 在每个表格后添加空行，以便更好地分隔
                md_file.write('\n\n')

        insert_logger.info(f"转换完成。Markdown 文件已保存为 {markdown_file}")
        return markdown_file

    @staticmethod
    def load_text(file_path):
        encodings = ['utf-8', 'iso-8859-1', 'windows-1252']

        for encoding in encodings:
            try:
                loader = TextLoader(file_path, encoding=encoding)
                docs = loader.load()
                insert_logger.info(f"TextLoader {encoding} success: {file_path}")
                return docs
            except Exception:
                insert_logger.warning(f"TextLoader {encoding} error: {file_path}, {traceback.format_exc()}")

        insert_logger.error(f"Failed to load file with all attempted encodings: {file_path}")
        return []

    @staticmethod
    def copy_images(image_root_path, output_dir):
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        # 获取当前目录下所有jpg文件
        images = [f for f in os.listdir(image_root_path) if f.endswith('.jpg')]
        # 复制到指定目录
        for image in images:
            single_image_path = os.path.join(image_root_path, image)
            insert_logger.info(f"copy image: {single_image_path} -> {output_dir}")
            shutil.copy(single_image_path, output_dir)

    @get_time
    async def split_file_to_docs(self):
        insert_logger.info(f"start split file to docs, file_path: {self.file_name}")
        if self.faq_dict:
            docs = [Document(page_content=self.faq_dict['question'], metadata={"faq_dict": self.faq_dict})]
        elif self.file_url:
            insert_logger.info("load url: {}".format(self.file_url))
            docs = await self.url_to_documents(self.file_path, self.file_name, self.file_url)
            if docs is None:
                try:
                    article = newspaper.article(self.file_url, timeout=120)
                    docs = [
                        Document(page_content=article.text, metadata={"title": article.title, "url": self.file_url})]
                except Exception as e:
                    insert_logger.error(f"newspaper get url error: {self.file_url}, {traceback.format_exc()}")
                    loader = MyRecursiveUrlLoader(url=self.file_url)
                    docs = loader.load()
        elif self.file_path.lower().endswith(".md"):
            try:
                docs = convert_markdown_to_langchaindoc(self.file_path)
                docs = self.markdown_process(docs)
            except Exception as e:
                insert_logger.error(
                    f"convert_markdown_to_langchaindoc error: {self.file_path}, {traceback.format_exc()}")
                loader = UnstructuredFileLoader(self.file_path, strategy="fast")
                docs = loader.load()
        elif self.file_path.lower().endswith(".txt"):
            docs = self.load_text(self.file_path)
        elif self.file_path.lower().endswith(".pdf"):
            try:
                data = {
                    'filename': self.file_path,
                    'save_dir': os.path.dirname(self.file_path)
                }
                headers = {"content-type": "application/json"}
                response = requests.post(f"http://{LOCAL_PDF_PARSER_SERVICE_URL}/pdfparser",json=data, headers=headers)
                response_json = response.json()
                markdown_file = response_json.get('markdown_file')
                docs = convert_markdown_to_langchaindoc(markdown_file)
                docs = self.markdown_process(docs)
                images_dir = os.path.join(IMAGES_ROOT_PATH, self.file_id)
                self.copy_images(os.path.dirname(markdown_file), images_dir)
            except Exception as e:
                insert_logger.warning(
                    f'Error in Powerful PDF parsing: {traceback.format_exc()}, use fast PDF parser instead.')
                loader = UnstructuredPaddlePDFLoader(self.file_path, strategy="fast")
                docs = loader.load()
        elif self.file_path.lower().endswith(".jpg") or self.file_path.lower().endswith(
                ".png") or self.file_path.lower().endswith(".jpeg"):
            txt_file_path = self.image_ocr_txt(filepath=self.file_path)
            loader = TextLoader(txt_file_path, autodetect_encoding=True)
            docs = loader.load()
        elif self.file_path.lower().endswith(".docx"):
            try:
                loader = UnstructuredWordDocumentLoader(self.file_path, strategy="fast")
                docs = loader.load()
            except Exception as e:
                insert_logger.warning('Error in Powerful Word parsing, use docx2txt instead.')
                text = docx2txt.process(self.file_path)
                docs = [Document(page_content=text)]
        elif self.file_path.lower().endswith(".xlsx"):
            try:
                markdown_file = self.excel_to_markdown(self.file_path, os.path.dirname(self.file_path))
                docs = convert_markdown_to_langchaindoc(markdown_file)
                docs = self.markdown_process(docs)
            except Exception as e:
                insert_logger.warning('Error in Powerful Excel parsing, use openpyxl instead.')
                docs = []
                excel_file = pd.ExcelFile(self.file_path)
                sheet_names = excel_file.sheet_names
                for idx, sheet_name in enumerate(sheet_names):
                    xlsx = pd.read_excel(self.file_path, sheet_name=sheet_name, engine='openpyxl')
                    csv_file_path = self.file_path[:-5] + f'_{idx}.csv'
                    xlsx.to_csv(csv_file_path, index=False)
                    insert_logger.info('xlsx2csv: %s', csv_file_path)
                    loader = CSVLoader(csv_file_path, autodetect_encoding=True,
                                       csv_args={"delimiter": ",", "quotechar": '"'})
                    docs.extend(loader.load())
        elif self.file_path.lower().endswith(".pptx"):
            loader = UnstructuredPowerPointLoader(self.file_path, strategy="fast")
            docs = loader.load()
        elif self.file_path.lower().endswith(".eml"):
            loader = UnstructuredEmailLoader(self.file_path, strategy="fast")
            docs = loader.load()
        elif self.file_path.lower().endswith(".csv"):
            loader = CSVLoader(self.file_path, autodetect_encoding=True, csv_args={"delimiter": ",", "quotechar": '"'})
            docs = loader.load()
        else:
            raise TypeError("文件类型不支持，目前仅支持：[md,txt,pdf,jpg,png,jpeg,docx,xlsx,pptx,eml,csv]")
        self.inject_metadata(docs)

    def inject_metadata(self, docs: List[Document]):
        # 这里给每个docs片段的metadata里注入file_id
        new_docs = []
        for doc in docs:
            page_content = re.sub(r'[\n\t]+', '\n', doc.page_content).strip()
            new_doc = Document(page_content=page_content)
            new_doc.metadata["user_id"] = self.user_id
            new_doc.metadata["kb_id"] = self.kb_id
            new_doc.metadata["file_id"] = self.file_id
            new_doc.metadata["file_name"] = self.file_name
            new_doc.metadata["nos_key"] = self.file_location
            new_doc.metadata["file_url"] = self.file_url
            new_doc.metadata["title_lst"] = doc.metadata.get("title_lst", [])
            new_doc.metadata["has_table"] = doc.metadata.get("has_table", False)
            # 从文本中提取图片数量：![figure]（x-figure-x.jpg）
            new_doc.metadata["images"] = re.findall(r'!\[figure]\(\d+-figure-\d+.jpg.*?\)', page_content)
            new_doc.metadata["page_id"] = doc.metadata.get("page_id", 0)
            kb_name = self.mysql_client.get_knowledge_base_name([self.kb_id])[0][2]
            metadata_infos = {"知识库名": kb_name, '文件名': self.file_name}
            new_doc.metadata['headers'] = metadata_infos

            if 'faq_dict' not in doc.metadata:
                new_doc.metadata['faq_dict'] = {}
            else:
                new_doc.metadata['faq_dict'] = doc.metadata['faq_dict']
            new_docs.append(new_doc)
        if new_docs:
            insert_logger.info('langchain analysis content head: %s', new_docs[0].page_content[:100])
        else:
            insert_logger.info('langchain analysis docs is empty!')

        # merge short docs
        insert_logger.info(f"before merge doc lens: {len(new_docs)}")
        child_chunk_size = min(DEFAULT_CHILD_CHUNK_SIZE, int(self.chunk_size / 2))
        merged_docs = []
        for doc in new_docs:
            if not merged_docs:
                merged_docs.append(doc)
            else:
                last_doc = merged_docs[-1]
                if num_tokens_embed(last_doc.page_content) + num_tokens_embed(doc.page_content) <= child_chunk_size or \
                        num_tokens_embed(doc.page_content) < child_chunk_size / 4:
                    tmp_content = doc.page_content
                    print(last_doc.metadata['title_lst'], tmp_content)
                    for title in last_doc.metadata['title_lst']:
                        tmp_content = tmp_content.replace(title, '')
                    last_doc.page_content += '\n\n' + tmp_content
                    # for title in last_doc.metadata['title_lst']:
                    #     last_doc.page_content = self.remove_substring_after_first(last_doc.page_content, '![figure]')
                    last_doc.metadata['title_lst'] += doc.metadata.get('title_lst', [])
                    last_doc.metadata['has_table'] = last_doc.metadata.get('has_table', False) or doc.metadata.get(
                        'has_table', False)
                    last_doc.metadata['images'] += doc.metadata.get('images', [])
                else:
                    merged_docs.append(doc)
        insert_logger.info(f"after merge doc lens: {len(merged_docs)}")
        self.docs = merged_docs
